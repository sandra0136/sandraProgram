#!/usr/local/bin python3
# -*- coding: utf-8 -*-
from openpyxl import Workbook
wb = Workbook()
ws = wb.active
# ws['A1'] = 42
# 在新的一列中添加
# ws.append([1, 2, 3])
# import datetime
# ws['A3'] = datetime.datetime.now()
hight = [180, 160, 170, 155]
weight = [90, 60, 70, 50]
ws['A1'] = 'heigh'
ws['B1'] = 'weigh'
for i in range(len(hight)):
    ws.cell(row=2+i, column=1, value=hight[i])
    ws.cell(row=2+i, column=2, value=weight[i])
wb.save("samle.xlsx")