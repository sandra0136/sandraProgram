#!/usr/local/bin python3
# -*- coding: utf-8 -*-
from openpyxl import Workbook,load_workbook

class PracticeExcel:

    def create(self):
        wb = Workbook()
        ws = wb.active
        ws['A1'] = '身高'
        ws['B1'] = '体重'
        self.heigt = [180, 160, 170, 155]
        self.weigh = [90, 60, 80, 50]
        for i in range(len(self.heigt)):
            ws.cell(2+i, 1, value=self.heigt[i])
            ws.cell(2+i, 2, value=self.weigh[i])
        wb.save('sample2.xlsx')


    def health_weight(self):
        ld = load_workbook(filename='sample2.xlsx')
        sheet = ld.active
        sheet['C1'] = "备注"
        for i in range(len(self.heigt)):
            height = sheet.cell(row=2+i, column=1).value
            weight = sheet.cell(row=2+i, column=2).value
            health_w = (height - 70)*0.6
            if weight != health_w:
                print("这是健康体重", weight)
                sheet.cell(row=2+i, column=3).value = "健康体重"
        ld.save("sample3.xlsx")



p = PracticeExcel()
p.create()
p.health_weight()